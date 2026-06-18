from openpyxl import load_workbook
import os

files = [
    r'C:\Users\tmdgu\Downloads\7조_API_명세서_최종.xlsx',
    r'C:\Users\tmdgu\Downloads\7조_요구사항 정의서_진짜최종 (1).xlsx'
]

target = '이승혁'

for file_path in files:
    print('\n' + '='*80)
    print('FILE: ' + os.path.basename(file_path))
    print('='*80 + '\n')
    
    if not os.path.exists(file_path):
        print('File not found: ' + file_path + '\n')
        continue
    
    try:
        wb = load_workbook(file_path)
        print('Sheets: ' + str(wb.sheetnames) + '\n')
        
        found_any = False
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                has_target = False
                for cell in row:
                    if cell and target in str(cell):
                        has_target = True
                        break
                
                if has_target:
                    found_any = True
                    print('>>> SHEET: ' + sheet_name + ' | ROW ' + str(row_idx))
                    for col_idx, cell in enumerate(row, 1):
                        print('  Col ' + str(col_idx) + ': ' + str(cell))
                    print()
        
        if not found_any:
            print('No rows containing "' + target + '" found.\n')
            
    except Exception as e:
        print('Error: ' + str(type(e).__name__) + ': ' + str(e) + '\n')
