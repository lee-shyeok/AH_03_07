# -*- coding: utf-8 -*-
import sys
import os
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import load_workbook

files = [
    (r'C:\Users\tmdgu\Downloads\7조_요구사항 정의서_진짜최종.xlsx', '요구사항'),
    (r'C:\Users\tmdgu\Downloads\7조_API_명세서_최종.xlsx', 'API'),
]

target = '이승혁'

for file_path, label in files:
    print(f'\n{"="*80}')
    print(f'FILE: {label} ({os.path.basename(file_path)})')
    print('='*80)

    if not os.path.exists(file_path):
        print(f'  NOT FOUND: {file_path}')
        continue

    wb = load_workbook(file_path, read_only=True)
    print(f'Sheets: {wb.sheetnames}')

    for sh in wb.sheetnames:
        ws = wb[sh]
        headers = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = row
                continue
            row_str = ' | '.join(str(c) for c in row if c is not None)
            if target in row_str:
                print(f'\n  [시트: {sh} / 행 {i+1}]')
                if headers:
                    for h, v in zip(headers, row):
                        if v is not None:
                            print(f'    {h}: {v}')
                else:
                    print(f'    {row}')
