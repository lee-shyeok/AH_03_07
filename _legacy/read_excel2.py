# -*- coding: utf-8 -*-
import sys
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import load_workbook

# 1. 요구사항 중 이승혁 담당
req_file = r'C:\Users\tmdgu\Downloads\7조_요구사항 정의서_진짜최종.xlsx'
wb = load_workbook(req_file, read_only=True)
ws = wb['요구사항 정의서']
rows = list(ws.iter_rows(values_only=True))
headers = rows[0]
col = {h: i for i, h in enumerate(headers) if h}

print("=== 이승혁 담당 요구사항 ===")
for r in rows[1:]:
    if r[col.get('담당자',99)] and '이승혁' in str(r[col.get('담당자',99)]):
        print(f"  {r[col['요구사항 ID']]} | {r[col['Category']]} | {r[col.get('Depth1','')]} / {r[col.get('Depth2','')]} | 우선순위:{r[col['우선순위']]} | 상태:{r[col.get('상태','')]}")

# 2. API 명세서 전체 엔드포인트 출력
api_file = r'C:\Users\tmdgu\Downloads\7조_API_명세서_최종.xlsx'
wb2 = load_workbook(api_file, read_only=True)
print("\n=== API 시트 목록 ===", wb2.sheetnames)

for shname in wb2.sheetnames:
    ws2 = wb2[shname]
    rows2 = list(ws2.iter_rows(values_only=True))
    if not rows2:
        continue
    print(f"\n--- 시트: {shname} ---")
    for r in rows2[:80]:
        if any(c for c in r if c is not None):
            print('  ' + ' | '.join(str(c) if c is not None else '' for c in r[:8]))
